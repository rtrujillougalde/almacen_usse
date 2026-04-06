"""
p_reportes.py - Página de Reportes (UI con Streamlit)

Contiene la interfaz de usuario para generar reportes PDF de entradas y salidas,
con filtros por fecha y centro de costo.
La lógica de acceso a datos se delega a data.py y la generación de reportes se
resuelve en este módulo.
"""

from io import BytesIO
from xml.sax.saxutils import escape

import streamlit as st
import pandas as pd
import datetime
from openpyxl.styles import Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table as ExcelTable, TableStyleInfo

from reportlab.lib.pagesizes import letter
from reportlab.lib.units import inch
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle

from data import (
    get_all_cc,
    get_report_data,
    
    get_comparacion_data,
)
from pdf_styles import (
    PDF_CC_COLOR,
    PDF_CC_FONT_SIZE,
    PDF_COL_WIDTHS,
    PDF_COL_WIDTHS_COMP,
    PDF_COL_WIDTHS_METADATA,
    draw_pdf_page_number,
    PDF_LOGO_FIRST_PAGE_HEIGHT,
    PDF_LOGO_FIRST_PAGE_WIDTH,
    PDF_LOGO_FIRST_PAGE_Y_OFFSET,
    PDF_LOGO_LATER_PAGE_HEIGHT,
    PDF_LOGO_LATER_PAGE_WIDTH,
    PDF_LOGO_LATER_PAGE_Y_OFFSET,
    PDF_LOGO_X,
    PDF_TABLE_STYLE_COMPARACION,
    PDF_TABLE_STYLE_REPORTE,
    PDF_MATERIAL_FONT_SIZE,
    PDF_MATERIAL_LEADING,
    PDF_SUMMARY_COLOR,
    PDF_SUMMARY_FONT_SIZE,
    PDF_TABLE_METADATA,
    PDF_TITLE_COLOR,
    PDF_TITLE_FONT_SIZE,
    PDF_ENTRADAS_TOTAL_COLOR,
    PDF_ENTRADAS_TOTAL_FONT_SIZE,
    PDF_TABLE_TEXT_COLOR,
)
from utils import LOGO_PATH


# =============================================================================
# HELPERS DE UI
# =============================================================================

def _draw_logo(canvas, doc, width, height, y_offset):
    canvas.saveState()
    canvas.drawImage(
        LOGO_PATH,
        x=PDF_LOGO_X,
        y=doc.pagesize[1] - y_offset,
        width=width,
        height=height,
        preserveAspectRatio=True,
        mask="auto",
    )
    canvas.restoreState()


def _draw_first_page_decorators(canvas, doc):
    _draw_logo(
        canvas,
        doc,
        PDF_LOGO_FIRST_PAGE_WIDTH,
        PDF_LOGO_FIRST_PAGE_HEIGHT,
        PDF_LOGO_FIRST_PAGE_Y_OFFSET,
    )


def _draw_later_page_decorators(canvas, doc):
    _draw_logo(
        canvas,
        doc,
        PDF_LOGO_LATER_PAGE_WIDTH,
        PDF_LOGO_LATER_PAGE_HEIGHT,
        PDF_LOGO_LATER_PAGE_Y_OFFSET,
    )
    draw_pdf_page_number(canvas, doc)


def _get_material_cell_style(styles):
    """Retorna el estilo común para celdas de la columna Material en PDFs."""
    return ParagraphStyle(
        "MaterialCell",
        parent=styles["BodyText"],
        fontSize=PDF_MATERIAL_FONT_SIZE,
        leading=PDF_MATERIAL_LEADING,
        alignment=1,
        textColor=PDF_TABLE_TEXT_COLOR,
    )


def _build_metadata_table(generated_at, cc_value):
    meta_table_data = [
        ["Fecha", "Hora", "C.C"],
        [
            generated_at.strftime("%d/%m/%Y"),
            generated_at.strftime("%H:%M:%S"),
            str(cc_value) if cc_value else "N/A",
        ],
    ]
    meta_table = Table(
        meta_table_data,
        colWidths=PDF_COL_WIDTHS_METADATA,
        hAlign="CENTER",
    )
    meta_table.setStyle(TableStyle(PDF_TABLE_METADATA))
    return meta_table


def _format_excel_worksheet(
    worksheet,
    currency_headers=None,
    convert_to_table=True,
    table_name=None,
):
    centered_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    horizontal_border = Side(style="thin", color="D3D3D3")
    no_vertical_border = Side(style=None)
    row_only_border = Border(
        left=no_vertical_border,
        right=no_vertical_border,
        top=horizontal_border,
        bottom=horizontal_border,
    )
    currency_headers = set(currency_headers or [])
    currency_format = '[$$-es-MX]#,##0.00'

    header_row = next(worksheet.iter_rows(min_row=1, max_row=1))
    currency_column_indexes = {
        cell.column for cell in header_row if cell.value in currency_headers
    }

    material_col_index = None
    for cell in header_row:
        if cell.value == "Material":
            material_col_index = cell.column
            break

    if material_col_index is not None:
        # Aproximacion de pixeles a ancho Excel (325 px ~ 46)
        worksheet.column_dimensions[get_column_letter(material_col_index)].width = 46

    for row in worksheet.iter_rows():
        for cell in row:
            cell.alignment = centered_alignment
            cell.border = row_only_border

            if cell.row > 1 and cell.column in currency_column_indexes and cell.value is not None:
                cell.number_format = currency_format

    if convert_to_table and worksheet.max_row >= 1 and worksheet.max_column >= 1:
        safe_table_name = table_name or f"Tabla_{worksheet.title}"
        safe_table_name = "".join(c if c.isalnum() else "_" for c in safe_table_name)
        if safe_table_name and safe_table_name[0].isdigit():
            safe_table_name = f"T_{safe_table_name}"

        table_ref = f"A1:{get_column_letter(worksheet.max_column)}{worksheet.max_row}"
        excel_table = ExcelTable(displayName=safe_table_name, ref=table_ref)
        excel_table.tableStyleInfo = TableStyleInfo(
            name="TableStyleMedium2",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False,
        )
        worksheet.add_table(excel_table)


def create_comparacion_pdf(comparacion_data):
    """
    Genera un PDF comparativo de entradas vs salidas por centro de costo.

    Args:
        comparacion_data (list[dict]): Datos obtenidos con get_comparacion_data().

    Returns:
        BytesIO: Buffer con el contenido del PDF generado.
    """
    buffer = BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=letter,
        rightMargin=40,
        leftMargin=40,
        topMargin=72,
        bottomMargin=18,
    )

    elements = []
    styles = getSampleStyleSheet()

    title_style = ParagraphStyle(
        "CustomTitle",
        parent=styles["Heading1"],
        fontSize=PDF_TITLE_FONT_SIZE,
        textColor=PDF_TITLE_COLOR,
        spaceAfter=20,
        alignment=1,
    )
    elements.append(Paragraph("Reporte Comparativo: Entradas vs Salidas", title_style))
    elements.append(Spacer(1, 0.2 * inch))

    generated_at = datetime.datetime.now()
    metadata_cc = None
    if comparacion_data:
        ccs_unicos = sorted({item["c_c"] for item in comparacion_data})
        metadata_cc = ccs_unicos[0] if len(ccs_unicos) == 1 else "Varios"
    elements.append(_build_metadata_table(generated_at, metadata_cc))
    elements.append(Spacer(1, 0.3 * inch))

    if comparacion_data:
        material_cell_style = _get_material_cell_style(styles)
        table_data = [[
            "Material",
            "Tipo",
            "Unidad",
            "Precio Unit.",
            "Salidas",
            "Entradas",
            "Usado",
            "Costo Mat. Usado",
        ]]

        for item in sorted(
            comparacion_data,
            key=lambda row: (str(row["c_c"]), str(row["material"])),
        ):
            table_data.append([
                Paragraph(escape(item["material"]), material_cell_style),
                item["tipo"],
                item["unidad_medida"],
                f"${(item['precio_unitario'] or 0):,.2f}",
                str(item["total_salida"]),
                str(item["total_entrada"]),
                str(item["usado"]),
                f"${item['costo_material_usado']:,.2f}",
            ])

        table = Table(table_data, colWidths=PDF_COL_WIDTHS_COMP)
        table.setStyle(TableStyle(PDF_TABLE_STYLE_COMPARACION))
        elements.append(table)
        elements.append(Spacer(1, 0.3 * inch))

        total_costo = sum(i["costo_material_usado"] for i in comparacion_data)
        summary_style = ParagraphStyle(
            "SummaryStyle",
            parent=styles["Normal"],
            fontSize=PDF_SUMMARY_FONT_SIZE,
            textColor=PDF_SUMMARY_COLOR,
            alignment=2,
        )
        elements.append(
            Paragraph(
                f"<b>Costo Total:</b> ${total_costo:,.2f}",
                summary_style,
            )
        )
    else:
        elements.append(
            Paragraph("No hay datos comparativos para mostrar.", styles["Normal"])
        )

    doc.build(
        elements,
        onFirstPage=_draw_first_page_decorators,
        onLaterPages=_draw_later_page_decorators,
    )
    buffer.seek(0)
    return buffer


def create_reporte_excel(report_data, movement_type="entrada"):
    """
    Genera un archivo Excel con los datos de entradas o salidas.

    Args:
        report_data (list[tuple]): Datos del movimiento obtenidos desde la capa de consultas.
        movement_type (str): Tipo de movimiento, "entrada" o "salida".

    Returns:
        BytesIO: Buffer con el contenido del Excel generado.
    """
    rows = []
    sheet_name = "Entradas" if movement_type == "entrada" else "Salidas"

    for row in report_data:
        fecha_hora, cc, material, cantidad, precio_unitario, unidad = row
        total = cantidad * (precio_unitario or 0)
        rows.append({
            "Fecha/Hora": fecha_hora.strftime("%Y-%m-%d %H:%M:%S") if fecha_hora else "",
            "C.C": cc,
            "Material": material,
            "Cantidad": cantidad,
            "Unidad": unidad,
            "Precio Unit.": precio_unitario or 0,
            "Total": total,
        })

    df = pd.DataFrame(rows)
    buffer = BytesIO()
    with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name=sheet_name)
        _format_excel_worksheet(
            writer.sheets[sheet_name],
            currency_headers={"Precio Unit.", "Total"},
            table_name=f"Tabla_{sheet_name}",
        )
    buffer.seek(0)
    return buffer


def create_comparacion_excel(comparacion_data):
    """
    Genera un archivo Excel comparativo de entradas vs salidas.

    Args:
        comparacion_data (list[dict]): Datos obtenidos con get_comparacion_data().

    Returns:
        BytesIO: Buffer con el contenido del Excel generado.
    """
    df = pd.DataFrame(comparacion_data)
    df.columns = [
        "C.C", "Material", "Tipo", "Unidad", "Precio Unit.",
        "Entradas", "Salidas",
        "Usado", "Costo Mat. Usado",
    ]
    buffer = BytesIO()
    with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Comparativo")
        _format_excel_worksheet(
            writer.sheets["Comparativo"],
            currency_headers={"Precio Unit.", "Costo Mat. Usado"},
            table_name="Tabla_Comparativo",
        )
    buffer.seek(0)
    return buffer


def create_reporte_pdf(report_data, movement_type="entrada"):
    """
    Genera un PDF con los datos de entradas o salidas.

    Args:
        report_data (list[tuple]): Datos del movimiento obtenidos desde la capa de consultas.
        movement_type (str): Tipo de movimiento, "entrada" o "salida".

    Returns:
        BytesIO: Buffer con el contenido del PDF generado.
    """
    buffer = BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=letter,
        rightMargin=50,
        leftMargin=50,
        topMargin=72,
        bottomMargin=18,
    )

    elements = []
    styles = getSampleStyleSheet()

    title_style = ParagraphStyle(
        "CustomTitle",
        parent=styles["Heading1"],
        fontSize=PDF_TITLE_FONT_SIZE,
        textColor=PDF_TITLE_COLOR,
        spaceAfter=20,
        alignment=1,
    )
    report_label = "Entradas" if movement_type == "entrada" else "Salidas"
    elements.append(Paragraph(f"Reporte de {report_label} de Almacén", title_style))
    elements.append(Spacer(1, 0.2 * inch))

    generated_at = datetime.datetime.now()
    elements.append(
        _build_metadata_table(
            generated_at,
            report_data[0][1] if report_data else None,
        )
    )
    elements.append(Spacer(1, 0.3 * inch))

    if report_data:
        material_cell_style = _get_material_cell_style(styles)
        
        table_data = [
            ["Fecha/Hora", "Material", "Cantidad", "Unidad", "Precio Unit.", "Total"]
        ]
        total_general = 0

        for row in report_data:
            fecha_hora, cc, material, cantidad, precio_unitario, unidad =row
            total = cantidad * (precio_unitario or 0)
            total_general += total
            table_data.append([
                fecha_hora.strftime("%Y-%m-%d %H:%M") if fecha_hora else "",
                Paragraph(escape(material), material_cell_style),
                str(cantidad),
                str(unidad),
                f"${precio_unitario:.2f}" if precio_unitario else "$0.00",
                f"${total:.2f}",
            ])

        table = Table(table_data, colWidths=PDF_COL_WIDTHS)
        table.setStyle(TableStyle(PDF_TABLE_STYLE_REPORTE))
        elements.append(table)
        elements.append(Spacer(1, 0.3 * inch))

        total_style = ParagraphStyle(
            "TotalStyle",
            parent=styles["Normal"],
            fontSize=PDF_ENTRADAS_TOTAL_FONT_SIZE,
            textColor=PDF_ENTRADAS_TOTAL_COLOR,
            alignment=2,
        )
        elements.append(
            Paragraph(f"<b>TOTAL GENERAL:</b> ${total_general:,.2f}", total_style)
        )
    else:
        elements.append(
            Paragraph(f"No hay datos de {movement_type}s para mostrar.", styles["Normal"])
        )

    doc.build(
        elements,
        onFirstPage=_draw_first_page_decorators,
        onLaterPages=_draw_later_page_decorators,
    )
    buffer.seek(0)
    return buffer

def render_filter_section(key_prefix=""):
    """
    Renderiza la sección de filtros (centro de costo y opcionalmente fechas).

    Args:
        key_prefix (str): Prefijo para las claves de los widgets (evita conflictos).

    Returns:
        tuple: (fecha_inicio_str, fecha_fin_str, cc_selected,
                fecha_inicio, fecha_fin, cc_seleccionados)
    """
    

    filtrar_fechas = st.checkbox(
        "Filtrar por rango de fechas",
        value=False,
        key=f"{key_prefix}filtrar_fechas",
    )

    fecha_inicio = None
    fecha_fin = None
    fecha_inicio_str = None
    fecha_fin_str = None

    if filtrar_fechas:
        col1, col2 = st.columns(2)
        with col1:
            fecha_inicio = st.date_input(
                "Fecha de inicio",
                value=datetime.date(datetime.datetime.now().year, 1, 1),
                key=f"{key_prefix}fecha_inicio",
            )
        with col2:
            fecha_fin = st.date_input(
                "Fecha de fin",
                value=datetime.date.today(),
                key=f"{key_prefix}fecha_fin",
            )
        fecha_inicio_str = fecha_inicio.strftime("%Y-%m-%d")
        fecha_fin_str = fecha_fin.strftime("%Y-%m-%d")

    

    return (
        fecha_inicio_str,
        fecha_fin_str,
        fecha_inicio,
        fecha_fin
    )


def build_report_filename(report_type, fecha_inicio, fecha_fin, cc_seleccionados, extension="pdf"):
    """
    Construye el nombre del archivo con el rango de fechas y centros de costo.

    Args:
        report_type (str): Tipo de reporte ('entradas', 'salidas', 'comparativo').
        fecha_inicio: Objeto date de inicio.
        fecha_fin: Objeto date de fin.
        cc_seleccionados (int | str): Centro de costo seleccionado.
        extension (str): Extensión del archivo ('pdf' o 'xlsx').

    Returns:
        str: Nombre del archivo formateado.
    """
    cc_str = f"_cc_{cc_seleccionados}" if cc_seleccionados else ""
    fecha_str = (
        f"_{fecha_inicio.strftime('%d%m%Y')}_a_{fecha_fin.strftime('%d%m%Y')}"
        if fecha_inicio and fecha_fin
        else ""
    )
    return f"reporte_{report_type}{fecha_str}{cc_str}.{extension}"


def create_preview_report_df(entradas_data, movement_type="entrada"):
    """
    Crea un DataFrame de vista previa para entradas o salidas.

    Args:
        entradas_data (list[tuple]): Datos del reporte.
        movement_type (str): Tipo de movimiento, "entrada" o "salida".

    Returns:
        pd.DataFrame: DataFrame formateado para visualización.
    """
    preview_df = []
    for entrada in entradas_data:
        fecha_hora, cc, material, cantidad, precio_unitario, unidad = entrada
        total = cantidad * (precio_unitario or 0)
        preview_df.append({
            "Fecha/Hora": fecha_hora,
            "C.C": cc,
            "Material": material,
            "Cantidad": cantidad,
            "Unidad": unidad,
            "Precio Unit.": f"${precio_unitario:.2f}" if precio_unitario else "$0.00",
            "Total": f"${total:.2f}",
        })
    return pd.DataFrame(preview_df)


def display_report_results(report_data, pdf_buffer, excel_buffer, pdf_name, excel_name, movement_type="entrada"):
    """
    Muestra descargas (PDF/Excel) y vista previa para entradas o salidas.
    """
    movement_label = "entrada" if movement_type == "entrada" else "salida"
    key_suffix = "entradas" if movement_type == "entrada" else "salidas"

    st.success(f"Se encontraron {len(report_data)} registros de {movement_label}")

    col1, col2 = st.columns(2)
    with col1:
        st.download_button(
            label="📄 Descargar PDF",
            data=pdf_buffer,
            file_name=pdf_name,
            mime="application/pdf",
            key=f"download_{key_suffix}_pdf",
        )
    with col2:
        st.download_button(
            label="📊 Descargar Excel",
            data=excel_buffer,
            file_name=excel_name,
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            key=f"download_{key_suffix}_excel",
        )

    st.subheader("Vista previa de datos")
    preview_df = create_preview_report_df(report_data, movement_type=movement_type)
    st.dataframe(preview_df, width='stretch')


# =============================================================================
# FUNCIONES PRINCIPALES DE PÁGINA
# =============================================================================

def reporte_main(
    fecha_inicio_str,
    fecha_fin_str,
    cc_selected,
    fecha_inicio,
    fecha_fin,
    can_generate,
    movement_type="entrada",
):
    """
    Función principal del reporte de entradas o salidas.
    Muestra filtros de fecha/CC, genera el PDF y ofrece descarga.
    """
    is_entrada = movement_type == "entrada"
    movement_label = "Entradas" if is_entrada else "Salidas"
    movement_label_singular = "entrada" if is_entrada else "salida"
    
    button_key = "entradas_pdf" if is_entrada else "salidas_pdf"

    st.title(f"📄 Reporte de {movement_label}")

    st.info(
        f"Este reporte extrae las {movement_type}s registradas de la base de datos, "
        "mostrando fecha, centro de costos, materiales, cantidades y precios."
    )

    if st.button(f"Generar Reporte de {movement_label}", key=button_key, disabled=not can_generate):
        with st.spinner(f"Obteniendo datos de {movement_type}s..."):
            try:
                report_data = get_report_data(
                    fecha_inicio_str, fecha_fin_str, cc_selected=cc_selected, movement_type=movement_type
                )
            except Exception as e:
                st.error(f"Error al obtener datos: {e}")
                return

        if report_data:
            pdf_buffer = create_reporte_pdf(report_data, movement_type=movement_type)
            excel_buffer = create_reporte_excel(report_data, movement_type=movement_type)
            pdf_name = build_report_filename(
                movement_type, fecha_inicio, fecha_fin, cc_selected, "pdf"
            )
            excel_name = build_report_filename(
                movement_type, fecha_inicio, fecha_fin, cc_selected, "xlsx"
            )
            display_report_results(
                report_data,
                pdf_buffer,
                excel_buffer,
                pdf_name,
                excel_name,
                movement_type=movement_type,
            )
        else:
            st.warning(
                f"No se encontraron registros de {movement_label_singular} en el rango de fechas especificado."
            )




def reporte_comparacion_main(
    fecha_inicio_str,
    fecha_fin_str,
    cc_selected,
    fecha_inicio,
    fecha_fin,
    can_generate,
):
    """
    Función principal del reporte comparativo entradas vs salidas.
    Muestra filtros de fecha/CC, genera el PDF y ofrece descarga.
    """
    st.title("📊 Reporte Comparativo")
    st.subheader("Entradas vs Salidas por Centro de Costo")
    st.info(
        "Este reporte compara los materiales y herramientas que entraron "
        "contra los que salieron para un mismo centro de costos, "
        "mostrando el porcentaje de uso real."
    )

    if st.button("Generar Reporte Comparativo", key="comparacion_pdf", disabled=not can_generate):
        with st.spinner("Obteniendo datos comparativos..."):
            try:
                comparacion_data = get_comparacion_data(
                    fecha_inicio_str, fecha_fin_str, cc_selected=cc_selected
                )
            except Exception as e:
                st.error(f"Error al obtener datos para comparar : {e}")
                return

        if comparacion_data:
            pdf_buffer = create_comparacion_pdf(comparacion_data)
            excel_buffer = create_comparacion_excel(comparacion_data)
            pdf_name = build_report_filename(
                "comparativo", fecha_inicio, fecha_fin, cc_selected, "pdf"
            )
            excel_name = build_report_filename(
                "comparativo", fecha_inicio, fecha_fin, cc_selected, "xlsx"
            )

            st.success(
                f"Se encontraron {len(comparacion_data)} materiales/herramientas para comparar"
            )

            col1, col2 = st.columns(2)
            with col1:
                st.download_button(
                    label="📄 Descargar PDF",
                    data=pdf_buffer,
                    file_name=pdf_name,
                    mime="application/pdf",
                    key="download_comparacion_pdf",
                )
            with col2:
                st.download_button(
                    label="📊 Descargar Excel",
                    data=excel_buffer,
                    file_name=excel_name,
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    key="download_comparacion_excel",
                )

            st.subheader("Vista previa de datos")
            preview_df = pd.DataFrame(comparacion_data)
            preview_df.columns = [
                "C.C", "Material", "Tipo", "Unidad", "Precio Unit.",
                "Total Entradas", "Total Salidas",
                "Usado", "Costo material usado",
            ]
            st.dataframe(preview_df, width='stretch')

            total_costo = sum(i["costo_material_usado"] for i in comparacion_data)
            st.metric("Costo Total de material usado", f"${total_costo:,.2f}")
        else:
            st.warning(
                "No se encontraron registros en el rango de fechas especificado."
            )

def main():
    key_prefix_map = {
        "Entradas": "entradas_",
        "Salidas": "salidas_",
        "Comparativo": "comparacion_",
    }
    col1, col2 = st.columns([1, 2])
    with col1:

        tipo_reporte = st.selectbox(
        "Selecciona el tipo de reporte",
        options=["Entradas", "Salidas", "Comparativo"],
        key="tipo_reporte_selector",
        )
    with col2:
        ccs = get_all_cc()

        cc_selected = st.selectbox(
        "Centro de Costos (CC)",
        options=ccs,
        index=None,
        placeholder="Selecciona un CC...",
        key=f"{key_prefix_map[tipo_reporte]}cc_selected",
    )
    

    (
        fecha_inicio_str,
        fecha_fin_str,
        fecha_inicio,
        fecha_fin,
    ) = render_filter_section(key_prefix=key_prefix_map[tipo_reporte])

    can_generate = bool(cc_selected)
    if not can_generate:
        st.warning("Debes seleccionar un Centro de Costos para generar el reporte.")

    if tipo_reporte == "Entradas":
        reporte_main(
            fecha_inicio_str,
            fecha_fin_str,
            cc_selected,
            fecha_inicio,
            fecha_fin,
            can_generate,
            movement_type="entrada",
        )
    elif tipo_reporte == "Salidas":
        reporte_main(
            fecha_inicio_str,
            fecha_fin_str,
            cc_selected,
            fecha_inicio,
            fecha_fin,
            can_generate,
            movement_type="salida",
        )
    elif tipo_reporte == "Comparativo":
        reporte_comparacion_main(
            fecha_inicio_str,
            fecha_fin_str,
            cc_selected,
            fecha_inicio,
            fecha_fin,
            can_generate,
        )